"""Конвертация загруженных файлов — изображения, PDF, Excel."""

import base64
import io
import logging
import zipfile
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from PIL import Image

from log_utils import sanitize_for_log

logger = logging.getLogger(__name__)

# xlsx это zip с XML внутри, поэтому пара сотен килобайт разворачиваются в десятки мегабайт.
# Потолок текста стоит не ради памяти, а ради промпта — 500 тыс. символов это ~125 тыс. токенов.
MAX_EXCEL_UNPACKED_MB = 5
MAX_EXCEL_TEXT_CHARS = 500_000

# В файле данные сжаты, в памяти растр по три байта на пиксель, поэтому вес файла ни о чём
# не говорит. 25 Мпикс это скан A4 до 500 dpi. На качество не влияет, всё принятое
# уменьшается до MAX_IMAGE_SIDE.
MAX_IMAGE_PIXELS = 25_000_000

# Сторона, до которой уменьшаем картинку перед отправкой в модель. Совпадает с тем,
# до чего масштабирует сам API при detail=high, поэтому отправлять больше бесполезно.
MAX_IMAGE_SIDE = 2048


class FileParseError(ValueError):
    """Файл не разобрать. Несёт ключ локализации и параметры, имя файла подставляет вызывающий код."""

    def __init__(self, key: str, **params: object) -> None:
        self.key = key
        self.params = params
        super().__init__(key)


class ImageTooLargeError(FileParseError):
    """Изображение превышает потолок по числу пикселей."""


# Сюда попадает и переименованный в .xlsx файл другого формата — openpyxl читает только zip
_UNREADABLE_EXCEL = "serverError.excelUnreadable"

# Один ключ на оба потолка таблицы — действие пользователя одно, числа в логе сервера.
_EXCEL_TOO_BIG = "serverError.excelTooBig"


def _ensure_excel_unpacked_size(excel_bytes: bytes) -> None:
    """Отклоняет zip-бомбу по распакованному размеру архива, ничего не распаковывая.

    Размеры записей лежат в оглавлении архива, распаковывать для проверки нечего.
    Считаются ВСЕ записи, а не только листы — openpyxl при read_only читает ещё тему
    оформления и стили, поэтому бомба в theme1.xml иначе прошла бы мимо потолка.

    Raises:
        FileParseError: файл не является zip-архивом либо распакованное
            содержимое больше потолка.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(excel_bytes)) as zf:
            unpacked = sum(info.file_size for info in zf.infolist())
    except BadZipFile as e:
        raise FileParseError(_UNREADABLE_EXCEL) from e

    limit = MAX_EXCEL_UNPACKED_MB * 1024 * 1024
    if unpacked > limit:
        # Ни объём, ни потолок наружу не отдаём — по ним тривиально подбирается файл под границу
        logger.warning(
            "Excel отклонён: архив разворачивается в %.1f МБ при лимите %d МБ",
            unpacked / 1024 / 1024, MAX_EXCEL_UNPACKED_MB,
        )
        raise FileParseError(_EXCEL_TOO_BIG)


def image_to_base64(img: Image.Image, max_size: int = MAX_IMAGE_SIDE) -> str:
    """Конвертирует PIL.Image в base64-строку PNG.

    Уменьшает изображение если оно больше max_size по любой стороне,
    чтобы не превышать лимиты LLM на размер изображений.
    Не мутирует входной объект — работает с копией при необходимости.

    Args:
        img: исходное изображение.
        max_size: максимальный размер стороны в пикселях.

    Returns:
        Base64-строка PNG-изображения.
    """
    if img.width > max_size or img.height > max_size:
        img = img.copy()
        img.thumbnail((max_size, max_size), Image.LANCZOS)

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode("utf-8")


def excel_to_text(excel_bytes: bytes) -> str:
    """Конвертирует Excel-файл в текстовое представление таблицы.

    Формирует читаемую таблицу с ячейками, разделёнными ``|``.

    Args:
        excel_bytes: содержимое Excel-файла в байтах.

    Returns:
        Текстовое представление всех листов Excel-файла.

    Raises:
        FileParseError: файл не читается как xlsx либо выходит за лимит
            распакованного архива или объёма текста.
    """
    _ensure_excel_unpacked_size(excel_bytes)

    try:
        wb = load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException) as e:
        raise FileParseError(_UNREADABLE_EXCEL) from e

    parts: list[str] = []
    chars_seen = 0

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_written = False

            # Читаем потоком, а не list(...) — xlsx умеет объявить миллионы строк
            for row in ws.iter_rows(values_only=True):
                if not header_written:
                    header = f"=== Sheet: {sheet_name} ==="
                    chars_seen += len(header) + 1
                    parts.append(header)
                    header_written = True

                cells = [str(cell) if cell is not None else "" for cell in row]
                # Длина считается ДО join — склейка аллоцирует, проверять после неё поздно.
                # Потолок распакованного архива тут не помогает, общий текст хранится в нём однажды.
                row_chars = sum(len(c) for c in cells) + 3 * max(0, len(cells) - 1)
                if chars_seen + row_chars > MAX_EXCEL_TEXT_CHARS:
                    logger.warning(
                        "Excel отклонён: текст перевалил за %d символов на листе «%s»",
                        MAX_EXCEL_TEXT_CHARS, sanitize_for_log(str(sheet_name)),
                    )
                    raise FileParseError(_EXCEL_TOO_BIG)

                chars_seen += row_chars + 1
                parts.append(" | ".join(cells))

            if header_written:
                parts.append("")  # пустая строка между листами
    finally:
        wb.close()

    return "\n".join(parts)


def open_image(image_bytes: bytes) -> Image.Image:
    """Открывает изображение и отклоняет слишком большое по числу пикселей.

    `Image.open` читает только заголовок, поэтому размер проверяется до декодирования.
    Уменьшение до `MAX_IMAGE_SIDE` делается здесь же — открытые картинки лежат списком
    до сборки запроса, и полный растр держался бы всё это время на каждый файл.

    Raises:
        ImageTooLargeError: пикселей больше `MAX_IMAGE_PIXELS`. Несёт разрешение картинки.
    """
    img = Image.open(io.BytesIO(image_bytes))
    pixels = img.width * img.height
    if pixels > MAX_IMAGE_PIXELS:
        # В сообщении только размер самой картинки, порог наружу не отдаём
        logger.warning(
            "Изображение отклонено: %d×%d = %.1f Мпикс при лимите %d Мпикс",
            img.width, img.height, pixels / 1_000_000, MAX_IMAGE_PIXELS // 1_000_000,
        )
        raise ImageTooLargeError(
            "serverError.imageTooLarge", width=img.width, height=img.height,
        )
    img.load()  # PIL ленив, битые данные вылезают здесь, а не на open()
    img.thumbnail((MAX_IMAGE_SIDE, MAX_IMAGE_SIDE), Image.LANCZOS)
    return img


def is_image_file(filename: str) -> bool:
    """Проверяет, является ли файл изображением по расширению.

    Args:
        filename: имя файла.

    Returns:
        True если расширение соответствует изображению.
    """
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    return ext in ("png", "jpg", "jpeg", "webp", "bmp")


def is_pdf_file(filename: str) -> bool:
    """Проверяет, является ли файл PDF.

    Args:
        filename: имя файла.

    Returns:
        True если расширение ``.pdf``.
    """
    return filename.lower().endswith(".pdf")


def is_excel_file(filename: str) -> bool:
    """Проверяет, является ли файл Excel.

    Args:
        filename: имя файла.

    Returns:
        True если расширение ``.xlsx`` или ``.xls``.
    """
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    return ext in ("xlsx", "xls")
