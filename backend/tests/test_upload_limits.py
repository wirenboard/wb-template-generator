"""Потолки на разбор присланных файлов.

Воркер uvicorn один, поэтому разбор подконтрольного пользователю файла обязан быть
ограничен по объёму работы.
"""

import io
import sys
import time
import tracemalloc
import zipfile
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from PIL import Image

# Добавляем backend/ в sys.path
sys.path.insert(0, str(Path(__file__).parent.parent))

import file_converter  # noqa: E402, I001
import main  # noqa: E402
from file_converter import (  # noqa: E402
    FileParseError,
    ImageTooLargeError,
    excel_to_text,
    open_image,
)
from user_errors import render  # noqa: E402


def _xlsx(rows: int, cols: int = 3, text: str = "value") -> bytes:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    for _ in range(rows):
        ws.append([text] * cols)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _png(width: int, height: int) -> bytes:
    buf = io.BytesIO()
    Image.new("RGB", (width, height), (7, 7, 7)).save(buf, format="PNG")
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

class TestExcelLimits:
    """Разбор таблицы ограничен объёмом листов и объёмом получаемого текста."""

    def test_normal_table_parsed(self):
        """Таблица штатного размера разбирается полностью."""
        text = excel_to_text(_xlsx(140, 5))

        assert text.count("\n") >= 140
        assert "value" in text

    def test_text_limit_enforced(self):
        """Текста больше лимита — отказ, даже если строк немного.

        Текст уходит в промпт, поэтому раздутая таблица это и счёт за токены, и риск,
        что запрос не примет провайдер. Срабатывает именно потолок текста, до потолка
        архива тут далеко.
        """
        cell = "x" * 20_000  # 3 ячейки в строке → ~60 КБ на строку
        with pytest.raises(FileParseError) as exc:
            excel_to_text(_xlsx(20, cols=3, text=cell))

        assert exc.value.key == "serverError.excelTooBig"

    def test_long_row_rejected_before_it_is_assembled(self):
        """Одна строка листа отсекается до склейки, а не после.

        Повторяющийся текст лежит в архиве однажды, а ссылок на него в строке сколько
        угодно, поэтому потолок распакованного объёма такую строку не ограничивает.
        Проверка после склейки означала бы, что память уже занята.
        """
        data = _xlsx(rows=1, cols=400, text="y" * 200_000)

        assert len(data) < 500 * 1024, "текст обязан лежать в архиве однажды"

        tracemalloc.start()
        try:
            with pytest.raises(FileParseError) as exc:
                excel_to_text(data)
            _current, peak = tracemalloc.get_traced_memory()
        finally:
            tracemalloc.stop()

        assert exc.value.key == "serverError.excelTooBig"
        assert peak < 10 * 1024 * 1024, f"строка всё-таки собиралась целиком: пик {peak} байт"

    def test_sheets_are_separated(self):
        """Заголовок листа появляется только у листа с данными.

        Текст уходит в промпт, поэтому лишний или потерянный заголовок тихо ухудшает анализ.
        """
        wb = Workbook()
        first = wb.active
        first.title = "Registers"
        first.append(["addr", "name"])
        wb.create_sheet("Empty")
        params = wb.create_sheet("Params")
        params.append(["baud", "9600"])
        buf = io.BytesIO()
        wb.save(buf)

        text = excel_to_text(buf.getvalue())

        assert "=== Sheet: Registers ===" in text
        assert "=== Sheet: Params ===" in text
        assert "=== Sheet: Empty ===" not in text

    def test_zip_bomb_rejected_before_opening(self):
        """Объём листов отсекается по оглавлению архива, до открытия книги.

        Главный потолок по времени — `load_workbook` пропорционален распакованному объёму.
        """
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("xl/worksheets/sheet1.xml", b"0" * (12 * 1024 * 1024))
        bomb = buf.getvalue()

        assert len(bomb) < 200 * 1024, "бомба должна быть маленькой в сжатом виде"

        t = time.monotonic()
        with pytest.raises(FileParseError) as exc:
            excel_to_text(bomb)

        assert exc.value.key == "serverError.excelTooBig"
        assert time.monotonic() - t < 0.5, "отказ обязан быть мгновенным"

    # Дозапись темы и стилей даёт дубликат имени, сумма всё равно считается по всем записям
    @pytest.mark.filterwarnings("ignore:Duplicate name")
    @pytest.mark.parametrize("part", [
        # openpyxl при read_only читает тему и стили целиком
        "xl/theme/theme1.xml",
        "xl/styles.xml",
        # Вложения openpyxl не читает, но исключений в подсчёте нет
        "xl/media/image1.png",
    ])
    def test_bomb_in_any_part_rejected(self, part):
        """Бомба в любой части архива отклоняется, не только в листах."""
        buf = io.BytesIO(_xlsx(10, 3))
        with zipfile.ZipFile(buf, "a", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(part, b"0" * (12 * 1024 * 1024))

        with pytest.raises(FileParseError) as exc:
            excel_to_text(buf.getvalue())

        assert exc.value.key == "serverError.excelTooBig"

    def test_old_xls_format_explained(self):
        """Старый формат объясняется текстом, а не общей ошибкой сервера."""
        ole_magic = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 512

        with pytest.raises(FileParseError) as exc:
            excel_to_text(ole_magic)

        # Ключ ведёт на текст с предложением сохранить как .xlsx
        assert exc.value.key == "serverError.excelUnreadable"
        assert ".xlsx" in render("serverError.excelUnreadable", {"file": "t.xls"})


# ---------------------------------------------------------------------------
# Изображения
# ---------------------------------------------------------------------------

class TestImageLimits:
    """Число пикселей ограничено до декодирования."""

    def test_normal_image_opened(self):
        """Обычная картинка открывается и декодируется."""
        img = open_image(_png(800, 600))

        assert img.size == (800, 600)

    def test_large_image_downscaled_on_open(self):
        """Крупная картинка уменьшается сразу, а не перед отправкой в модель."""
        img = open_image(_png(4000, 3000))

        assert max(img.size) == file_converter.MAX_IMAGE_SIDE
        assert img.size == (2048, 1536)

    def test_pixel_limit_enforced(self, monkeypatch):
        """Превышение потолка пикселей — отказ до декодирования."""
        # Потолок опускаем, чтобы не генерировать в тесте картинку на 40 Мпикс
        monkeypatch.setattr(file_converter, "MAX_IMAGE_PIXELS", 100_000)

        with pytest.raises(ImageTooLargeError) as exc:
            open_image(_png(500, 500))

        # Разрешение картинки уходит параметрами, порог наружу не отдаём
        assert exc.value.key == "serverError.imageTooLarge"
        assert exc.value.params == {"width": 500, "height": 500}

    def test_broken_image_raises(self):
        """Битые данные поднимают ошибку Pillow, а не проходят молча."""
        with pytest.raises(Exception) as exc:
            open_image(b"not an image at all")

        assert not isinstance(exc.value, ImageTooLargeError)


# ---------------------------------------------------------------------------
# Эндпоинт анализа
# ---------------------------------------------------------------------------

class TestAnalyzeUploadLimits:
    """Ограничения на входе, до чтения файлов в память."""

    @pytest.fixture
    def client(self, monkeypatch):
        # Без lifespan, очереди для отказов не нужны. Адрес LLM не проверяем, иначе тест пойдёт в DNS
        async def _skip_url_check(url, allow_private=False):
            return None

        monkeypatch.setattr(main, "ensure_public_llm_url", _skip_url_check)
        main._rate_limit_store.clear()
        return TestClient(main.app)

    def test_unsupported_extension_rejected(self, client):
        """Расширение не из списка отклоняется до чтения файла."""
        resp = client.post(
            "/api/analyze",
            files=[("files", ("table.xls", b"\xd0\xcf\x11\xe0", "application/vnd.ms-excel"))],
            data={"llm_api_url": "https://api.provider.example/v1"},
        )

        assert resp.status_code == 400
        assert resp.json()["message_key"] == "serverError.unsupportedFormat"

    def test_too_many_files_rejected(self, client):
        """Число файлов ограничено — расход умножается на их количество."""
        settings = main.get_settings()
        files = [
            ("files", (f"page{i}.png", b"x", "image/png"))
            for i in range(settings.MAX_FILES + 1)
        ]

        resp = client.post(
            "/api/analyze", files=files,
            data={"llm_api_url": "https://api.provider.example/v1"},
        )

        assert resp.status_code == 400
        assert str(settings.MAX_FILES) in resp.json()["detail"]
        assert "PDF" in resp.json()["detail"]


class TestStatusExposesLimits:
    """Потолки входа отдаёт сервер — интерфейс отсекает по ним набор до отправки."""

    def test_status_carries_input_limits(self):
        settings = main.get_settings()

        payload = TestClient(main.app).get("/api/status").json()

        assert payload["max_files"] == settings.MAX_FILES
        assert payload["max_file_size_mb"] == settings.MAX_FILE_SIZE_MB
        assert sorted(payload["allowed_extensions"]) == sorted(main._ALLOWED_EXTENSIONS)


class TestImportUploadLimits:
    """Импорт шаблона читает файл целиком, поэтому ограничен так же, как анализ."""

    @pytest.fixture
    def client(self):
        return TestClient(main.app)

    def test_large_template_rejected(self, client):
        """Файл больше потолка отклоняется до разбора."""
        settings = main.get_settings()
        payload = b"{}" + b" " * (settings.MAX_FILE_SIZE_MB * 1024 * 1024)

        resp = client.post(
            "/api/import-template",
            files=[("file", ("huge.json", payload, "application/json"))],
        )

        assert resp.status_code == 413
        assert resp.json()["message_key"] == "serverError.fileTooLarge"

    def test_normal_template_still_imported(self, client):
        """Обычный шаблон проходит — потолок не мешает штатному импорту."""
        template = (
            b'{"device": {"name": "T", "id": "t", "channels": '
            b'[{"name": "V", "reg_type": "holding", "address": 1}]}}'
        )

        resp = client.post(
            "/api/import-template",
            files=[("file", ("t.json", template, "application/json"))],
        )

        assert resp.status_code == 200
        assert resp.json()["registers"][0]["name"] == "V"
